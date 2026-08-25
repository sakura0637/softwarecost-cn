# -*- coding: utf-8 -*-
"""
设备价格库 ETL：解析桌面 10 站「*设备台账*.xlsx」→ server/seed/device_prices_seed.json

结构与分类规则（2026-08-25 重新梳理，修正旧版大量漏读）：
1. 设备判定：单价(col6) 非 None 即真实设备；分类标题/小计行单价必空 → 不入设备。
   （旧版把整数序号当大纲行、漏掉海量设备，导致分类下为空 / 总数虚低）
2. 分类层级用「编号小数点深度」入栈：
   - X.0          → 顶层分类（工程监控 / 视频监视 / 计算机网络 / 通信 / 安全监测 / 实体环境 …）
   - X.M          → 子分类（硬件设备 / 软件 / 控制专网 / 通信传输 / 通信交换 …）
   - X.M.K        → 孙分类
   - 序号为空的分类标题 → 取「栈顶深度+1」
   遇到更浅层级时弹出深层，保证栈顶=当前最近分类。
3. 公式兜底：data_only 读不到值时（公式无缓存），用 data_only=False 的工作簿求值
   （同表 =E*G、跨表 Sheet!A1、SUM(range)），避免漏读。
4. subsite：
   - 普通管理处：第一个 sheet「全站设备明细/全站设备汇总」→ subsite='全站设备汇总'（汇总，不归到子站）；
     其余 sheet = 各子站名。
   - 总调中心：无子站，sheet 名即分类；跳过「全站设备明细」（其与各分类 sheet 设备重复，
     分类 sheet 含数量且更完整），其余 sheet → category=sheet名、subsite='全站设备汇总'。

输出字段：station, subsite, category, subcategory, name, unit, brand_model, qty, unit_price, total_price, remark
"""
import os, glob, re, json
import openpyxl
from openpyxl.utils import column_index_from_string as col_idx

BASE = r"C:\Users\lenovo\Desktop\价格汇总\现有汇总"
OUT  = r"D:\softwarecost\server\seed\device_prices_seed.json"

# ───────────────────────── 公式求值 ─────────────────────────
_frm_cache = {}  # filename -> data_only=False workbook（普通模式，仅公式兜底时懒加载）

def _ref_parts(cell_only):
    m = re.match(r'^([A-Z]+)(\d+)$', cell_only)
    if not m:
        return None
    return col_idx(m.group(1)), int(m.group(2))

def _cell_val(wb_frm, sheet, c, r, depth):
    if depth > 30:
        return None
    try:
        ws = wb_frm[sheet]
        v = ws.cell(row=r, column=c).value
    except Exception:
        return None
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str) and v.startswith('='):
        return _eval_formula(v, sheet, wb_frm, depth + 1)
    return None

def _eval_formula(formula, cur_sheet, wb_frm, depth=0):
    expr = formula[1:].strip()

    def sum_rep(m):
        tot = 0.0
        for part in m.group(1).split(','):
            part = part.strip()
            if ':' in part:
                a, b = part.split(':', 1)
                loc_a = _ref_parts(a.strip()) if '!' not in a else None
                # 解析 range 两端（可能带 sheet）
                sa = cur_sheet; sb = cur_sheet
                am = re.match(r'^(?:([^!]+?)!)?([A-Z]+\d+)$', a.strip())
                bm = re.match(r'^(?:([^!]+?)!)?([A-Z]+\d+)$', b.strip())
                if am and bm:
                    if am.group(1): sa = am.group(1).strip().strip("'")
                    if bm.group(1): sb = bm.group(1).strip().strip("'")
                    ca, ra = _ref_parts(am.group(2)); cb, rb = _ref_parts(bm.group(2))
                    if sa == sb and ca and cb:
                        r1, r2 = sorted([ra, rb]); c1, c2 = sorted([ca, cb])
                        for rr in range(r1, r2 + 1):
                            for cc in range(c1, c2 + 1):
                                v = _cell_val(wb_frm, sa, cc, rr, depth)
                                if v: tot += v
            else:
                rm = re.match(r'^(?:([^!]+?)!)?([A-Z]+\d+)$', part)
                if rm:
                    sh = rm.group(1).strip().strip("'") if rm.group(1) else cur_sheet
                    c, r = _ref_parts(rm.group(2))
                    v = _cell_val(wb_frm, sh, c, r, depth)
                    if v: tot += v
        return str(tot)

    expr = re.sub(r'SUM\(([^)]*)\)', sum_rep, expr, flags=re.I)

    def rep(m):
        sheet_part = m.group(1)  # 可能含 ! 的 sheet 前缀
        cell = m.group(2)
        sheet = cur_sheet
        if sheet_part:
            sheet = sheet_part.rstrip('!').strip().strip("'")
        c, r = _ref_parts(cell)
        if not c:
            return '0'
        v = _cell_val(wb_frm, sheet, c, r, depth)
        return str(v if v is not None else 0)

    expr = re.sub(r"((?:[^!()]+?)!)?([A-Z]+\d+)", rep, expr)
    try:
        return float(eval(expr, {'__builtins__': {}}, {}))
    except Exception:
        return None

def to_float(v):
    if v is None:
        return None
    if isinstance(v, str) and v.strip() == '':
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

def eval_cell_frm(fname, sheet, r, c):
    """data_only 读不到（公式无缓存）时，用 data_only=False 工作簿求值。"""
    if fname not in _frm_cache:
        _frm_cache[fname] = openpyxl.load_workbook(fname, data_only=False, read_only=False)
    wb_frm = _frm_cache[fname]
    try:
        f = wb_frm[sheet].cell(row=r, column=c).value
    except Exception:
        return None
    if isinstance(f, str) and f.startswith('='):
        return _eval_formula(f, sheet, wb_frm)
    return to_float(f)


# ───────────────────────── 解析 ─────────────────────────
def find_header(ws):
    for i in range(1, min(ws.max_row, 15) + 1):
        vals = [ws.cell(row=i, column=c).value for c in range(1, 9)]
        if any(isinstance(x, str) and x.strip() == '设备名称' for x in vals):
            return i
    return 1

def sn_depth(sn):
    if sn is None:
        return None
    s = str(sn).strip()
    if s in ('', '·', 'None'):
        return None
    if '.' in s:
        parts = s.split('.')
        if len(parts) == 2 and parts[1] == '0':
            return 2
        if len(parts) == 2:
            return 3
        return 4
    if s.isdigit():
        return 3
    return 3

SKIP = ('合计', '小计', '总计')

def parse_sheet(ws_val, fname, station, subsite, zongdiao, sheetname):
    hr = find_header(ws_val)
    rows = list(ws_val.iter_rows(min_row=hr, max_row=ws_val.max_row, max_col=8, values_only=True))
    stack = []  # (depth, name)
    devs = []
    missed = 0
    for idx, r in enumerate(rows):
        if not r or len(r) < 8:
            continue
        sn, name = r[0], r[1]
        if name is None:
            continue
        name = str(name).strip()
        if name == '设备名称' or name == '' or name == '·':
            continue
        if any(k in name for k in SKIP):
            continue
        real_row = hr + idx
        price = to_float(r[6])
        if price is None:
            qty_try = to_float(r[4])
            if qty_try is None:
                # 分类标题行
                d = sn_depth(sn)
                if d is None:
                    d = (stack[-1][0] + 1) if stack else 2
                while stack and stack[-1][0] >= d:
                    stack.pop()
                stack.append((d, name))
                continue
            price = eval_cell_frm(fname, sheetname, real_row, 7) or 0  # 有数量无单价→公式兜底
        # 设备
        if zongdiao and sheetname != '全站设备明细':
            cat = sheetname
            sub = stack[-1][1] if stack else None
        else:
            cat = stack[0][1] if stack else None
            if cat is None:
                cat = subsite  # 兜底：廊涿「光缆监测/安全监测设备」等独立汇总 sheet 无顶层分类标题
            sub = stack[-1][1] if (stack and (len(stack) > 1 or stack[-1][1] != cat)) else None
        qty = to_float(r[4]) or eval_cell_frm(fname, sheetname, real_row, 5)
        total = to_float(r[7]) or eval_cell_frm(fname, sheetname, real_row, 8)
        unit = r[2]
        brand = r[3]
        remark = r[5]
        devs.append({
            'station': station,
            'subsite': subsite,
            'category': cat,
            'subcategory': sub,
            'name': name,
            'unit': (str(unit).strip() if unit and str(unit).strip() not in ('·',) else None),
            'brand_model': (str(brand).strip() if brand and str(brand).strip() not in ('·',) else None),
            'qty': qty,
            'unit_price': price,
            'total_price': total,
            'remark': (str(remark).strip() if remark and str(remark).strip() not in ('·',) else None),
        })
    return devs


def norm_station(fn):
    s = fn.replace('设备台账', '').replace('.xlsx', '')
    s = s.replace('价格补全版', '').replace('极简版', '')  # 去版本描述词
    s = re.sub(r'[vV]\d+', '', s)                          # 去版本号 v1/v2/v8
    s = s.replace('_', '').strip()
    return s or fn


def main():
    files = sorted(glob.glob(os.path.join(BASE, '*设备台账*.xlsx')))
    files = [f for f in files if not os.path.basename(f).startswith('~$')]
    all_devs = []
    stat = {}
    for f in files:
        fn = os.path.basename(f)
        station = norm_station(fn)
        zongdiao = '总调中心' in fn
        wb_val = openpyxl.load_workbook(f, data_only=True, read_only=True)
        s_count = 0
        for sh in wb_val.sheetnames:
            if zongdiao and sh == '全站设备明细':
                continue  # 总调中心：分类 sheet 已含全部设备，跳过汇总避免重复
            # 总调中心无子站，subsite 统一为「全站设备汇总」（category 才是 sheet 名/分类）
            subsite = '全站设备汇总' if (zongdiao or ('全站' in sh)) else sh
            devs = parse_sheet(wb_val[sh], f, station, subsite, zongdiao, sh)
            all_devs.extend(devs)
            s_count += len(devs)
        stat[station] = s_count
        wb_val.close()
        print(f"  {station}: {s_count} 台/件")
    # 关闭公式缓存 wb
    for wb in _frm_cache.values():
        try: wb.close()
        except Exception: pass

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, 'w', encoding='utf-8') as fh:
        json.dump(all_devs, fh, ensure_ascii=False)
    print(f"\n总设备条数: {len(all_devs)}")
    print(f"已写出: {OUT}")
    # 分类空校验
    empty_cat = sum(1 for d in all_devs if not d['category'])
    print(f"category 为空: {empty_cat}")
    # 各站统计
    with open(OUT + '.stat.txt', 'w', encoding='utf-8') as fh:
        fh.write('\n'.join(f"{k}\t{v}" for k, v in stat.items()))


if __name__ == '__main__':
    main()
